from flask import Flask, render_template, request
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        num_events = int(request.form['num_events'])
        events_data = []

        for i in range(num_events):
            time_from = datetime.strptime(request.form[f'time_from_{i}'], '%H:%M')
            time_to = datetime.strptime(request.form[f'time_to_{i}'], '%H:%M')

            # Calculate the time duration in minutes
            time_duration = (time_to - time_from).seconds // 60

            program = request.form[f'program_{i}']
            subject = request.form[f'subject_{i}']
            presenter = request.form[f'presenter_{i}']
            events_data.append([i+1, time_from.strftime('%H:%M'), time_to.strftime('%H:%M'), time_duration, program, subject, presenter])

        date = request.form['date']  # Get the provided date
        title = f"SM Sabha Program Sunday {date}"

        # Generate the filename based on the title
        excel_file = f"{title}.xlsx"

        columns = ["Item", "From", "To", "Mins", "Programme", "Subject", "Presenter"]
        df = pd.DataFrame(events_data, columns=columns)

        wb = Workbook()
        ws = wb.active
        ws.title = "Program"

        # Create a font with specified properties
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')  # White font color

        # Merge and center cells for the title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply background color to the title cell (Orange Accent 2, lighter 40%)
        fill = PatternFill(start_color='FF9933', end_color='FF9933', fill_type='solid')  # Orange Accent 2, lighter 40%
        title_cell.fill = fill

        # Write the column headers
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Adjust column widths to fit content
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Adjust the width of the "Item" column
        ws.column_dimensions['A'].width = 6  # Set the width to 6 (adjust as needed)

        # Save the Excel file with the generated filename
        wb.save(excel_file)

        return f"Event program saved to {excel_file}"

    return render_template('index.html', num_events=1)
if __name__ == '__main__':
    app.run(debug=True)
